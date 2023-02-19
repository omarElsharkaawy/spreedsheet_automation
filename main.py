import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def calculate_new_price(filename):
    # load spreed sheet
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # loop starting from 2nd row to the max rows in sheet
    for row in range(2, sheet.max_row + 1):
        # detect cell
        cell = sheet.cell(row, 3)

        # calculate the new price
        new_price = cell.value * .9

        # select new cell to add new value
        new_price_cell = sheet.cell(row, 4)

        # passing new value to cell
        new_price_cell.value = new_price

    # select cells for chart data
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, max_col=4, min_col=4)

    # add data to chart
    chart = BarChart()
    chart.add_data(values)

    # add chart to sheet
    sheet.add_chart(chart, 'e2')

    # save changes in new file
    wb.save(filename)


calculate_new_price('transactions.xlsx')
